import time
import gspread as gspread
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from Auths import jira_login, jira_password
import httplib2
import apiclient
from oauth2client.service_account import ServiceAccountCredentials
import datetime


# from googleapiclient import discovery


def jira_commit():
    try:
        driver.get('https://fab-i-t.atlassian.net/browse/PROC-515?')
        time.sleep(2)
        driver.find_element(By.NAME, "username").send_keys(jira_login)
        time.sleep(1)
        driver.find_element(By.CLASS_NAME, "css-19r5em7").click()
        time.sleep(1)
        driver.find_element(By.NAME, "password").send_keys(jira_password)
        time.sleep(1)
        driver.find_element(By.ID, "login-submit").click()
        time.sleep(4)
        driver.find_element(By.CSS_SELECTOR, ".qbld2j-3.cxAoMh").click()
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, ".Input__InputElement-sc-1o6bj35-0.bfCuIo").send_keys('60m')
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, ".css-goggrm").click()
        time.sleep(3)
    except Exception as ex:
        print(ex)

    finally:
        driver.close()
        driver.quit()


def find_amount():
    driver.find_element(By.ID, "total").click()
    time.sleep(4)
    amount = driver.find_element(By.CLASS_NAME, "heading_title-text").text
    # print(amount)
    if "Учреждения" in amount:
        result = ''.join(amount.strip('Учреждения (количество: ').strip(')').split())
        # print(result)
        return result
    else:
        return ''.join(amount.strip('События (количество: ').strip(')').split())


def auth_pro_culture():
    gmail = 'Artem.Kluykovskiy@fabit.ru'
    password = 'Fakemask01'
    driver.get('https://pro.culture.ru/new/auth/login')
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "ant-input").send_keys(gmail)
    time.sleep(1)
    driver.find_elements(By.CLASS_NAME, 'ant-input')[1].send_keys(password)
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR,
                        ".ant-btn.ant-btn-primary").click()
    time.sleep(4)
    driver.find_element(By.CSS_SELECTOR, ".main-form_organization-link").click()
    time.sleep(3)


def institutions():
    categories_of_institution = (
        'teatry', 'muzei-i-galerei', 'koncertnye-ploshchadki', 'dvorcy-kultury-i-kluby', 'prochee',
        'obrazovatelnye-uchrezhdeniya', 'biblioteki', 'kinoteatry', 'kulturnoe-nasledie', 'pamyatniki', 'parki',
        'cirki'
    )
    terminal_status = ('true', 'false')
    global levels_of_subordination
    institution_values = dict()
    for status in terminal_status:
        for level in levels_of_subordination:
            sum_category_other = 0
            institution_values[f'{level}/{status}'] = []
            for i, one_category_of_institution in enumerate(categories_of_institution):
                if i <= 6:
                    driver.get(
                        f'https://pro.culture.ru/new/subordinate/organizations?subordinationLevel='
                        f'{level}&isPushkinsCard=true&hasTerminals={status}'
                        f'&category={one_category_of_institution}'
                    )
                    time.sleep(2.5)
                    try:
                        institution_values[f'{level}/{status}'].append(find_amount())
                    except Exception as ex:
                        institution_values[f'{level}/{status}'].append('0')
                else:
                    driver.get(
                        f'https://pro.culture.ru/new/subordinate/organizations?subordinationLevel='
                        f'{level}&isPushkinsCard=true&hasTerminals={status}'
                        f'&category={one_category_of_institution}'
                    )
                    try:
                        time.sleep(3)
                        sum_category_other += int(find_amount())
                    except Exception as ex:
                        pass
            institution_values[f'{level}/{status}'].append(str(sum_category_other))
            flag = True
            while flag:
                try:
                    driver.get(
                        f'https://pro.culture.ru/new/subordinate/organizations?isPushkinsCard=true&hasTerminals={status}&subordinationLevel={level}')
                    time.sleep(4)
                    all_org = find_amount()
                    flag = False
                except:
                    print(f"Упс, {status}/{level} не нашлось..попробуем еще раз! ")
                    flag = True
            # print(institution_values[f'{level}/{status}'])
            sum_org = sum_orgs(institution_values, level, status)
            # print(f"{level}/{status} = {all_org}")
            while all_org != str(sum_org):
                print(f'{level}/{status} не совпадает со значение из PRO',
                      f'\nПолученное значение = {institution_values[f"{level}/{status}"]}\nСумма - {sum_org}',
                      f'\nВсего дал сайт - {all_org}')
                institution_values[f'{level}/{status}'] = checking_inst(institution_values, status, level,
                                                                        categories_of_institution)
                sum_org = sum_orgs(institution_values, level, status)
            print(
                f"institution_values[f'{level}/{status}'] = {institution_values[f'{level}/{status}']}\nСумма = {sum_org}\nВсего дал сайт = {all_org}\nAll right!")
    return institution_values


def sum_orgs(institution_values, level, status):
    sum_org = 0
    for element in institution_values[f'{level}/{status}']:
        if element != '':
            sum_org += int(element)
    return sum_org


def checking_inst(institution_values, status, level, categories_of_institution):
    sum_category_other = 0
    # for value in institution_values[f'{level}/{status}']:
    for i, one_category_of_institution in enumerate(categories_of_institution):
        if i <= 6:
            try:
                driver.get(f'https://pro.culture.ru/new/subordinate/organizations?subordinationLevel='
                           f'{level}&isPushkinsCard=true&hasTerminals={status}'
                           f'&category={one_category_of_institution}')
                time.sleep(3)
                number = find_amount()
                if number != institution_values[f'{level}/{status}'][i]:
                    institution_values[f'{level}/{status}'][i] = number

            except:
                pass
        else:
            try:
                driver.get(
                    f'https://pro.culture.ru/new/subordinate/organizations?subordinationLevel='
                    f'{level}&isPushkinsCard=true&hasTerminals={status}'
                    f'&category={one_category_of_institution}'
                )

                time.sleep(2)
                number = find_amount()
                sum_category_other += int(number)
            except Exception as ex:
                pass
    if sum_category_other != institution_values[f'{level}/{status}'][-1]:
        institution_values[f'{level}/{status}'][-1] = sum_category_other

    return institution_values[f'{level}/{status}']


def events():
    global levels_of_subordination
    events_status = ('accepted', 'new')
    events_values = dict()
    for status in events_status:
        for level in levels_of_subordination:
            events_values[f"{status}/{level}"] = ''
            while events_values[f"{status}/{level}"] < '-1':
                try:
                    driver.get(
                        f"https://pro.culture.ru/new/subordinate/events?status={status}"
                        f"&isPushkinsCard=true&subordinationLevel={level}"
                    )
                    time.sleep(6)
                    events_values[f"{status}/{level}"] = find_amount()
                    # events_values[f"{status}/{level}"] = events_values[f"{status}/{level}"][-1]
                except Exception as ex:
                    # print(events_)
                    pass
            events_values[f"{status}/{level}"] = int(events_values[f"{status}/{level}"])
    pushka_status = ('accepted', 'new', 'rejected')
    for status_p in pushka_status:
        events_values[f'pushka/{status_p}'] = ''
        while events_values[f'pushka/{status_p}'] < '-1':
            try:
                driver.get(
                    f'https://pro.culture.ru/new/subordinate/events?'
                    f'status=accepted&isPushkinsCard=true&statusPushka={status_p}'
                )
                time.sleep(10)
                events_values[f'pushka/{status_p}'] = find_amount()
                time.sleep(4)
            # events_values[f'pushka/{status_p}'] = events_values[f'pushka/{status_p}'][-1]
            except Exception as ex:
                print(f"Не нашло события пушки: {status_p}")
        events_values[f'pushka/{status_p}'] = int(events_values[f'pushka/{status_p}'])
    events_values['all_pushka'] = '0'
  #  while events_values['all_pushka'] < '1':
   #   driver.get(
    #     'https://pro.culture.ru/new/subordinate/events?isPushkinsCard=true')  # Если что поменять местами с нижней строчкой
    
    #try:
     #  time.sleep(10)  
      # events_values['all_pushka'] = find_amount()
      # time.sleep(4)
    #except:
     #  print(f'All events of pushka  - Error')
     #if len(events_values['all_pushka']) != 1:
     #  events_values['all_pushka'] = events_values['all_pushka'][-1]
    events_values['all_pushka'] = int(events_values['all_pushka'])
    print(events_values)
    return events_values


levels_of_subordination = ('municipal', 'regional', 'federal')


def google_sheet_uploading():
    global inst
    global ev
    with open('result.txt', 'w') as file:
        file.write(str(inst))
        file.write(str(ev))
    CREDENTIALS_FILE = 'creds.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        CREDENTIALS_FILE,
        ['http  s://www.googleapis.com/auth/spreadsheets',
         'https://www.googleapis.com/auth/drive'])
    httpAuth = credentials.authorize(httplib2.Http())
    service = discovery.build('sheets', 'v4', http=httpAuth)
    sheet_id = 0
    try:
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                'valueInputOption': "USER_ENTERED",
                'data': [
                    {'range': "B3:I5",
                     'majorDimension': "ROWS",
                     'values': [inst['municipal/true'], inst['regional/true'], inst['federal/true']]},
                    {'range': "B11:I13",
                     'majorDimension': "ROWS",
                     'values': [inst['municipal/false'], inst['regional/false'], inst['federal/false']]},
                    {'range': "D28:D30",
                     'majorDimension': "ROWS",
                     'values': [[ev['accepted/municipal']], [ev['accepted/regional']], [ev['accepted/federal']]]},
                    {'range': "E28:E30",
                     'majorDimension': "ROWS",
                     'values': [[ev['new/municipal']], [ev['new/regional']], [ev['new/federal']]]},
                    # Добавить строковый тип к событиям

                    {'range': "D40:E40",
                     'majorDimension': "COLUMNS",
                     'values': [[ev['pushka/accepted']], [ev['pushka/new']]]},
                    {'range': "E44",
                     'majorDimension': "COLUMNS",
                     'values': [ev['pushka/rejected']]},
                    {'range': "E35",
                     'majorDimension': "COLUMNS",
                     'values': [ev['all_pushka']]}
                ]
            }
        ).execute()
        copy_sheet_to_another_spreadsheet_request_body = {
            # The ID of the spreadsheet to copy the sheet to.
            'destination_spreadsheet_id': '1gq0BR2l9dr8ucjubNCMn_7kLbGHnDHr_zvTerHC1wDQ',
        }

        service.spreadsheets().sheets().copyTo(spreadsheetId=spreadsheet_id, sheetId=sheet_id,
                                               body=copy_sheet_to_another_spreadsheet_request_body).execute()
    except:
        print("File was not sent into google sheets!")

    time.sleep(5)


def delete_rubish_sheet():
    gc = gspread.service_account(filename='creds.json')
    sh = gc.open_by_key('1WD7muBaVq8aXmy1ayTh_yp44J37mDPV5W0qTQ_CXsCk')
    worksheet_list = sh.worksheets()
    while len(worksheet_list) > 40:
        worksheet = sh.worksheet(worksheet_list[-1].title)
        sh.del_worksheet(worksheet)
        print(f"{worksheet} удалён !")
        # time.sleep(10)
        worksheet_list = sh.worksheets()


def rename_list():
    gc = gspread.service_account(filename='creds.json')
    sh = gc.open_by_key('1gq0BR2l9dr8ucjubNCMn_7kLbGHnDHr_zvTerHC1wDQ')
    worksheet_list = sh.worksheets()
    worksheet = sh.worksheet(worksheet_list[-1].title)
    worksheet.title = '.'.join(str(datetime.date.today()).strip('2022').split('-')[::-1]).strip('.')
    print(f"{worksheet.title}")


spreadsheet_id = '1fCM2MIHd0FmfmyGfh0oIEJCYBue1P1mI4XQ5vjHAGxA'


def add_in_cell(number, rows, ws):
    columns = ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I')
    for i, value in enumerate(inst[rows]):
        ws[f'{columns[i] + number}'] = int(value)


def excel_adding():
    global inst
    global ev

    source = r"D:\Codes\morning_reports\Reports_file\Пушкинская карта. Ежедневный отчет.xlsx"
    wb = openpyxl.load_workbook(source)  # path to the Excel file
    ws = wb.active

    for rows in inst:
        if 'true' in rows:
            if "municipal" in rows:
                add_in_cell('3', rows, ws)
            elif "reg" in rows:
                add_in_cell('4', rows, ws)
            elif "fed" in rows:
                add_in_cell('5', rows, ws)
        else:
            if "municipal" in rows:
                add_in_cell('11', rows, ws)
            elif "reg" in rows:
                add_in_cell('12', rows, ws)
            elif "fed" in rows:
                add_in_cell('13', rows, ws)
    global levels_of_subordination
    # D28:E30
    ws['D28'] = ev['accepted/municipal']
    ws['D29'] = ev['accepted/regional']
    ws['D30'] = ev['accepted/federal']
    ws['E28'] = ev['new/municipal']
    ws['E29'] = ev['new/regional']
    ws['E30'] = ev['new/federal']
    ws['E35'] = ev['all_pushka']
    ws['D40'] = ev['pushka/accepted']
    ws['E40'] = ev['pushka/new']
    ws['E44'] = ev['pushka/rejected']
    wb.save(source)
    write_results(inst, ev)


def write_results(inst, ev):
    with open('result.txt', 'w') as file:
        file.write(str(inst))
        file.write(str(ev))


def pushka_report_start():
    excel_adding()
    delete_rubish_sheet()


# rename_list()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)
auth_pro_culture()
inst = institutions()
ev = events()
pushka_report_start()
driver.close()
driver.quit()

# delete_rubish_sheet()
